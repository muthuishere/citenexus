"""XlsxExtractor — each sheet: a heading (its name) + ``col: value`` row blocks.

The csv twin for workbooks (§8): a sheet's first row is the schema (carried on
each row block's ``structure_path``), every subsequent row is a table block of
``col: value`` pairs, zip-shortest. ``page`` is the 1-based sheet index. Cell
values render with cross-language-stable rules (the Rust twin must match
byte-for-byte): empty → ``""``, bool → ``true``/``false``, integral numbers →
integer digits, everything else → its string form.
"""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from citenexus.extract.plain import open_binary
from citenexus.extract.types import (
    BlockKind,
    ExtractedBlock,
    ExtractedDoc,
    SourceType,
    StructureType,
)
from citenexus.plugins.base import ExtractorPlugin


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


class XlsxExtractor(ExtractorPlugin):
    """Sheets become heading-scoped runs of schema-zipped table blocks."""

    plugin_version = "xlsx/1"

    def __init__(self, document_id: str | None = None) -> None:
        self.document_id = document_id

    def extract(self, source: Any) -> ExtractedDoc:
        opened, doc_id, source_uri = open_binary(source, self.document_id)
        workbook = load_workbook(opened, read_only=True, data_only=True)

        blocks: list[ExtractedBlock] = []
        order = 0
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            blocks.append(
                ExtractedBlock(
                    order=order,
                    kind=BlockKind.heading,
                    text=sheet.title,
                    page=sheet_index,
                    level=1,
                )
            )
            order += 1

            rows = sheet.iter_rows(values_only=True)
            first = next(rows, None)
            if first is None:
                continue
            header_cells = [_cell_text(v) for v in first]
            while header_cells and not header_cells[-1]:
                header_cells.pop()
            header = tuple(header_cells)
            if not header:
                continue

            row_index = 0
            for row in rows:
                values = [_cell_text(v) for v in row]
                if not any(values):
                    continue
                pairs = list(zip(header, values, strict=False))  # zip-shortest
                rendered = ", ".join(f"{col}: {val}" for col, val in pairs)
                blocks.append(
                    ExtractedBlock(
                        order=order,
                        kind=BlockKind.table,
                        text=rendered,
                        page=sheet_index,
                        level=row_index,
                        structure_path=header,
                        cells=tuple(val for _, val in pairs),
                    )
                )
                order += 1
                row_index += 1

        has_rows = any(b.kind is BlockKind.table for b in blocks)
        return ExtractedDoc(
            document_id=doc_id,
            source_type=SourceType.xlsx,
            structure_type=StructureType.table_schema if has_rows else StructureType.none,
            source_uri=source_uri,
            blocks=tuple(blocks),
        )
