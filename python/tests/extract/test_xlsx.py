"""XlsxExtractor — per-sheet heading + ``col: value`` table blocks (csv twin)."""

from __future__ import annotations

import io

from openpyxl import Workbook

from citenexus.extract.types import BlockKind, SourceType, StructureType
from citenexus.extract.xlsx import XlsxExtractor


def _workbook_bytes() -> bytes:
    wb = Workbook()
    people = wb.active
    people.title = "People"
    people.append(["name", "age", "active"])
    people.append(["ada", 36, True])
    people.append(["alan", 41.5, False])
    scores = wb.create_sheet("Scores")
    scores.append(["team", "points"])
    scores.append(["red", 30])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_two_sheets_extract_sheet_scoped_rows() -> None:
    doc = XlsxExtractor(document_id="doc").extract(_workbook_bytes())
    assert doc.source_type is SourceType.xlsx
    assert doc.structure_type is StructureType.table_schema

    kinds = [b.kind for b in doc.blocks]
    assert kinds == [
        BlockKind.heading,
        BlockKind.table,
        BlockKind.table,
        BlockKind.heading,
        BlockKind.table,
    ]
    assert [b.order for b in doc.blocks] == [0, 1, 2, 3, 4]

    people_heading, ada, alan, scores_heading, red = doc.blocks
    assert people_heading.text == "People"
    assert people_heading.level == 1
    assert people_heading.page == 1
    assert ada.text == "name: ada, age: 36, active: true"
    assert ada.structure_path == ("name", "age", "active")
    assert ada.level == 0
    assert ada.page == 1
    assert alan.text == "name: alan, age: 41.5, active: false"
    assert alan.level == 1
    assert scores_heading.text == "Scores"
    assert scores_heading.page == 2
    assert red.text == "team: red, points: 30"
    assert red.page == 2


def test_empty_sheets_yield_headings_only() -> None:
    wb = Workbook()
    wb.active.title = "Empty"
    buf = io.BytesIO()
    wb.save(buf)
    doc = XlsxExtractor(document_id="doc").extract(buf.getvalue())
    assert [b.kind for b in doc.blocks] == [BlockKind.heading]
    assert doc.blocks[0].text == "Empty"
    assert doc.structure_type is StructureType.none


def test_rows_zip_shortest_against_header() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["a", "b"])
    ws.append(["only-a", "b-val", "spill"])
    buf = io.BytesIO()
    wb.save(buf)
    doc = XlsxExtractor(document_id="doc").extract(buf.getvalue())
    assert doc.blocks[1].text == "a: only-a, b: b-val"


def test_fully_empty_rows_are_skipped() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["a", "b"])
    ws.append([None, None])
    ws.append(["x", None])
    buf = io.BytesIO()
    wb.save(buf)
    doc = XlsxExtractor(document_id="doc").extract(buf.getvalue())
    texts = [b.text for b in doc.blocks if b.kind is BlockKind.table]
    assert texts == ["a: x, b: "]
